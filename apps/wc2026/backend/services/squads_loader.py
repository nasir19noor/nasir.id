"""
Load 26-player squads from the curated wall-chart spreadsheet.

The spreadsheet is the canonical source of *static* squad data only:
jersey, position, age, caps, club, captain. Tournament-goal counts
(`wc_goals`) are NOT read from here — they're recomputed from ESPN
match details on every hourly refresh (services/espn_fetcher.py).

Idempotent: re-running upserts rows by (team_id, player_name).
"""
import logging
import os
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database import SessionLocal
from models import Player, Team

logger = logging.getLogger(__name__)

DEFAULT_XLSX = "/app/World_Cup_2026_Wall_Chart.xlsx"

# Country-name aliases — sheet name → our Team.name (for the few that differ).
NAME_ALIASES: dict[str, str] = {
    # Add entries here if a sheet country name ever fails to match Team.name.
}


def _index_teams(db: Session) -> dict[str, Team]:
    """Build a case-insensitive name → Team index."""
    idx: dict[str, Team] = {}
    for t in db.query(Team).all():
        idx[t.name.upper()] = t
    return idx


def _resolve_team(idx: dict[str, Team], country: str) -> Team | None:
    key = NAME_ALIASES.get(country, country).upper()
    return idx.get(key)


def load_squads(xlsx_path: str | Path | None = None) -> dict:
    path = Path(xlsx_path or os.environ.get("SQUADS_XLSX", DEFAULT_XLSX))
    if not path.exists():
        logger.warning("Squads file not found: %s — skipping", path)
        return {"loaded": False, "reason": f"file missing: {path}"}

    logger.info("Loading squads from %s", path)
    wb = load_workbook(path, data_only=True)
    if "Squads" not in wb.sheetnames:
        return {"loaded": False, "reason": "no 'Squads' sheet in workbook"}

    ws = wb["Squads"]
    db = SessionLocal()
    upserted = 0
    skipped_countries: set[str] = set()
    try:
        teams_by_name = _index_teams(db)
        header_seen = False

        for row in ws.iter_rows(values_only=True):
            if not row or all(c is None for c in row):
                continue
            if not header_seen:
                if str(row[0] or "").strip() == "Group":
                    header_seen = True
                continue

            country = str(row[1] or "").strip()
            if not country or not row[3]:
                continue

            team = _resolve_team(teams_by_name, country)
            if team is None:
                skipped_countries.add(country)
                continue

            jersey     = int(row[2]) if isinstance(row[2], (int, float)) else None
            name       = str(row[3]).strip()
            is_captain = name.endswith("(C)")
            if is_captain:
                name = name[:-3].strip()
            pos      = (str(row[4] or "").strip().upper()[:3]) or None
            age      = int(row[5]) if isinstance(row[5], (int, float)) else None
            caps     = int(row[6]) if isinstance(row[6], (int, float)) else 0
            intl     = int(row[7]) if isinstance(row[7], (int, float)) else 0
            club     = (str(row[8] or "").strip()) or None
            # wc_goals is intentionally NOT read here — ESPN owns that field.

            existing = (db.query(Player)
                          .filter(Player.team_id == team.id, Player.name == name)
                          .one_or_none())
            if existing is None:
                existing = Player(team_id=team.id, name=name)  # wc_goals defaults to 0
                db.add(existing)
            existing.jersey_number = jersey
            existing.position      = pos
            existing.age           = age
            existing.caps          = caps
            existing.intl_goals    = intl
            existing.club          = club
            existing.is_captain    = is_captain
            upserted += 1

        db.commit()
        result = {
            "loaded":             True,
            "players_upserted":   upserted,
            "skipped_countries":  sorted(skipped_countries),
        }
        logger.info("Squads load: %s", result)
        return result
    except Exception as e:
        db.rollback()
        logger.exception("Squads load failed: %s", e)
        raise
    finally:
        db.close()
