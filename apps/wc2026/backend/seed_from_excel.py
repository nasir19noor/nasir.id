"""
Seed the database from the canonical wall-chart spreadsheet.

The spreadsheet (`World_Cup_2026_Wall_Chart.xlsx`) carries curated data that no
public API gives us — group layouts, fixture order, squads, captains, club
affiliations. ESPN later overlays live scores and kickoff times on top of these
rows.

Idempotent: safe to re-run; existing teams/players/fixtures are updated, not
duplicated.
"""
import os
import sys
import logging
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database import SessionLocal
from models import Team, Player, Fixture, KnockoutMatch

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# 3-letter code -> (full name, ISO 3166-1 alpha-2 for flagcdn).
# Codes pulled from the spreadsheet's Group Stage sheet.
TEAM_META = {
    "MEX": ("Mexico",            "mx"),
    "SAF": ("South Africa",      "za"),
    "KOR": ("South Korea",       "kr"),
    "CZE": ("Czech Republic",    "cz"),
    "CAN": ("Canada",            "ca"),
    "BOS": ("Bosnia and Herzegovina", "ba"),
    "QAT": ("Qatar",             "qa"),
    "SWI": ("Switzerland",       "ch"),
    "BRA": ("Brazil",            "br"),
    "MOR": ("Morocco",           "ma"),
    "HAI": ("Haiti",             "ht"),
    "SCO": ("Scotland",          "gb-sct"),
    "USA": ("United States",     "us"),
    "PAR": ("Paraguay",          "py"),
    "AUS": ("Australia",         "au"),
    "TUR": ("Turkey",            "tr"),
    "ARG": ("Argentina",         "ar"),
    "POR": ("Portugal",          "pt"),
    "ESP": ("Spain",              "es"),
    "FRA": ("France",             "fr"),
    "GER": ("Germany",            "de"),
    "BEL": ("Belgium",            "be"),
    "NED": ("Netherlands",        "nl"),
    "ENG": ("England",            "gb-eng"),
    "ITA": ("Italy",              "it"),
    "CRO": ("Croatia",            "hr"),
    "URU": ("Uruguay",            "uy"),
    "COL": ("Colombia",           "co"),
    "ECU": ("Ecuador",            "ec"),
    "JPN": ("Japan",              "jp"),
    "IRN": ("Iran",               "ir"),
    "EGY": ("Egypt",              "eg"),
    "GHA": ("Ghana",              "gh"),
    "NGA": ("Nigeria",            "ng"),
    "SEN": ("Senegal",            "sn"),
    "ALG": ("Algeria",            "dz"),
    "TUN": ("Tunisia",            "tn"),
    "CIV": ("Ivory Coast",        "ci"),
    "CRC": ("Costa Rica",         "cr"),
    "PAN": ("Panama",             "pa"),
    "JAM": ("Jamaica",            "jm"),
    "POL": ("Poland",             "pl"),
    "DEN": ("Denmark",            "dk"),
    "NOR": ("Norway",             "no"),
    "AUT": ("Austria",            "at"),
    "SRB": ("Serbia",             "rs"),
    "SVN": ("Slovenia",           "si"),
    "ALB": ("Albania",            "al"),
    "WAL": ("Wales",              "gb-wls"),
    "IRL": ("Ireland",            "ie"),
    "UKR": ("Ukraine",            "ua"),
    "SVK": ("Slovakia",           "sk"),
    "HUN": ("Hungary",            "hu"),
    "ROU": ("Romania",            "ro"),
}


def _ensure_team(db: Session, code: str, group_letter: str | None = None) -> Team:
    code = code.strip().upper()
    team = db.query(Team).filter(Team.code == code).one_or_none()
    name, iso2 = TEAM_META.get(code, (code, None))
    if team is None:
        team = Team(code=code, name=name, iso2=iso2, group_letter=group_letter)
        db.add(team)
        db.flush()
    else:
        if group_letter and team.group_letter != group_letter:
            team.group_letter = group_letter
        if iso2 and not team.iso2:
            team.iso2 = iso2
        if name and team.name == code:
            team.name = name
    return team


def _seed_groups(db: Session, ws) -> int:
    """Parse 'Group Stage' sheet. Returns number of fixtures upserted."""
    fixtures = 0
    current_group: str | None = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        first = str(row[0] or "").strip()
        if first.startswith("GROUP "):
            current_group = first.split()[1][:1]
            continue
        if not current_group:
            continue
        # Match rows: (match_no, home_code, home_score, away_score, away_code, ...)
        if isinstance(row[0], (int, float)) and row[1] and row[4]:
            match_no   = int(row[0])
            home_code  = str(row[1]).strip().upper()
            away_code  = str(row[4]).strip().upper()
            home_score = int(row[2]) if isinstance(row[2], (int, float)) else None
            away_score = int(row[3]) if isinstance(row[3], (int, float)) else None

            home = _ensure_team(db, home_code, current_group)
            away = _ensure_team(db, away_code, current_group)

            fx = (db.query(Fixture)
                    .filter(Fixture.group_letter == current_group,
                            Fixture.match_no     == match_no)
                    .one_or_none())
            if fx is None:
                fx = Fixture(group_letter=current_group, match_no=match_no,
                             home_team_id=home.id, away_team_id=away.id)
                db.add(fx)
            else:
                fx.home_team_id = home.id
                fx.away_team_id = away.id
            # Only seed scores when the sheet has both — don't overwrite live data later.
            if fx.home_score is None and home_score is not None:
                fx.home_score = home_score
                fx.away_score = away_score
                fx.status     = "finished"
            fixtures += 1
    return fixtures


def _seed_squads(db: Session, ws) -> int:
    """Parse 'Squads' sheet. Returns number of players upserted."""
    players = 0
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        if not header_seen:
            if str(row[0] or "").strip() == "Group":
                header_seen = True
            continue
        group_letter = str(row[0] or "").strip()[:1] if row[0] else None
        country      = str(row[1] or "").strip()
        if not country or not row[3]:
            continue
        # Match team by name (fallback by code if equal length).
        team = db.query(Team).filter(Team.name == country).one_or_none()
        if team is None:
            continue
        if group_letter and team.group_letter != group_letter:
            team.group_letter = group_letter

        jersey     = int(row[2]) if isinstance(row[2], (int, float)) else None
        name       = str(row[3]).strip()
        is_captain = name.endswith("(C)")
        if is_captain:
            name = name[:-3].strip()
        pos        = str(row[4] or "").strip().upper()[:3] or None
        age        = int(row[5]) if isinstance(row[5], (int, float)) else None
        caps       = int(row[6]) if isinstance(row[6], (int, float)) else 0
        intl       = int(row[7]) if isinstance(row[7], (int, float)) else 0
        club       = str(row[8] or "").strip() or None
        wc_goals   = int(row[9]) if isinstance(row[9], (int, float)) else 0

        existing = (db.query(Player)
                      .filter(Player.team_id == team.id, Player.name == name)
                      .one_or_none())
        if existing is None:
            existing = Player(team_id=team.id, name=name)
            db.add(existing)
        existing.jersey_number = jersey
        existing.position      = pos
        existing.age           = age
        existing.caps          = caps
        existing.intl_goals    = intl
        existing.club          = club
        existing.is_captain    = is_captain
        existing.wc_goals      = wc_goals
        players += 1
    return players


def _seed_knockout(db: Session, ws) -> int:
    """Parse 'Knockout Bracket' sheet — just the empty slot structure."""
    round_map = {
        "ROUND OF 32":  "r32",
        "ROUND OF 16":  "r16",
        "QUARTER-FINALS": "qf",
        "SEMI-FINALS":   "sf",
        "THIRD PLACE":   "third",
        "FINAL":         "final",
    }
    current_round: str | None = None
    slot = 0
    seeded = 0
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        first = str(row[0] or "").strip()
        if first.upper() in round_map:
            current_round = round_map[first.upper()]
            slot = 0
            continue
        if first.lower().startswith(("match", "quarter-final", "semi-final", "final", "third")) and current_round:
            slot += 1
            home_label = str(row[1] or "").strip() or None
            away_label = str(row[4] or "").strip() or None
            existing = (db.query(KnockoutMatch)
                          .filter(KnockoutMatch.round_code == current_round,
                                  KnockoutMatch.slot       == slot)
                          .one_or_none())
            if existing is None:
                existing = KnockoutMatch(round_code=current_round, slot=slot,
                                         home_label=home_label, away_label=away_label)
                db.add(existing)
            else:
                existing.home_label = home_label
                existing.away_label = away_label
            seeded += 1
    return seeded


def seed(xlsx_path: str | Path | None = None) -> dict:
    xlsx_path = Path(xlsx_path or os.environ.get(
        "SEED_XLSX",
        "/app/seed/World_Cup_2026_Wall_Chart.xlsx",
    ))
    if not xlsx_path.exists():
        logger.warning("Seed file not found: %s — skipping", xlsx_path)
        return {"seeded": False, "reason": "seed file missing"}

    logger.info("Seeding from %s", xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)

    db = SessionLocal()
    try:
        fixtures = _seed_groups(db,   wb["Group Stage"])
        players  = _seed_squads(db,   wb["Squads"])
        knockout = _seed_knockout(db, wb["Knockout Bracket"])
        db.commit()
        result = {
            "seeded":  True,
            "fixtures_upserted": fixtures,
            "players_upserted":  players,
            "knockout_slots":    knockout,
        }
        logger.info("Seed complete: %s", result)
        return result
    except Exception as e:
        db.rollback()
        logger.exception("Seed failed: %s", e)
        raise
    finally:
        db.close()


if __name__ == "__main__":
    print(seed(sys.argv[1] if len(sys.argv) > 1 else None))
